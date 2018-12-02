from email_function import email_function
from openpyxl import load_workbook

wb = load_workbook('email.xlsx')
sheet = wb.active

e = Email()

for row in sheet.iter_rows():
	name = row[0].value
	addr = row[1].value
	e.send_mail(name,addr) 