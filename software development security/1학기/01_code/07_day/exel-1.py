from openpyxl import load_workbook

# load_workbook 함수를 이용하여 엑셀 클래스 변수 생성
wb = load_workbook('test.xlsx')
# 활성화 된 시트를 sheet 변수로 설정
sheet = wb.active

sheet['A1'] = 'Number'
sheet['B1'] = 'Name'

sheet['A2'] = 1
sheet['B2'] = 'AAA'
sheet['A3'] = 2
sheet['B3'] = 'BBB'


# test02라는 파일이 있으면 덮어쓰기 됨
# 없으면 새로운 엑셀파일 생성
wb.save('test02.xlsx')
